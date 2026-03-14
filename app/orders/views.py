from django.shortcuts import render
import builtins
from ..orders.models import Order, OrderLine
from ..products.models import Product
from ..stocks.models import Stock
from .constants import ORDER_STATUS, ORDER_ROUTE_RECORD, ORDER_WOO_STATUS

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q, CharField, Count
from django.db.models.functions import Cast
from django.conf import settings
import json
import re
import logging

import csv
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from datetime import timedelta
from decimal import Decimal, InvalidOperation
from html import unescape
import re as py_re

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from ..orders.cron import push_order_to_wc, sync_woo_order_completed
from ..services.shippit_client import get_shipping_quote, build_parcel_attributes, normalize_parcel_attributes
from ..orders.cron import sync_wc_orders

logger = logging.getLogger(__name__)

def parse_bool(value):
	if isinstance(value, bool):
		return value
	if value is None:
		return False
	return str(value).strip().lower() in ['1', 'true', 'yes', 'on']

def _is_empty_delivery_date(value):
	if not value:
		return True
	return str(value).strip() == '1970-01-01'

def _clean_param(request, key):
	return (request.GET.get(key) or '').strip()

def _parse_decimal_value(value):
	if value is None:
		return None
	try:
		return Decimal(str(value).replace(',', '').strip())
	except (InvalidOperation, ValueError, TypeError):
		return None

def _apply_text_filter(queryset, field, raw):
	if not raw:
		return queryset
	return queryset.filter(**{f'{field}__icontains': raw})

def _apply_number_filter(queryset, field, raw):
	raw = (raw or '').strip()
	if not raw:
		return queryset

	op_match = re.match(r'^(<=|>=|<|>|=|!=)\s*(.+)$', raw)
	range_match = re.match(r'^(.+?)(?:\.\.|-)(.+)$', raw)

	if op_match:
		op, rhs_raw = op_match.groups()
		rhs = _parse_decimal_value(rhs_raw)
		if rhs is None:
			return queryset

		if op == '!=':
			return queryset.exclude(**{field: rhs})

		lookup_map = {
			'<=': 'lte',
			'>=': 'gte',
			'<': 'lt',
			'>': 'gt',
			'=': 'exact',
		}
		lookup = lookup_map.get(op)
		if not lookup:
			return queryset
		return queryset.filter(**{f'{field}__{lookup}': rhs})

	if range_match:
		left_raw, right_raw = range_match.groups()
		left = _parse_decimal_value(left_raw)
		right = _parse_decimal_value(right_raw)
		if left is None or right is None:
			return queryset
		low = min(left, right)
		high = max(left, right)
		return queryset.filter(**{f'{field}__gte': low, f'{field}__lte': high})

	value = _parse_decimal_value(raw)
	if value is None:
		return queryset
	return queryset.filter(**{field: value})

def _apply_date_filter(queryset, field, raw):
	raw = (raw or '').strip()
	if not raw:
		return queryset

	def parse_token(token):
		token = (token or '').strip()
		if re.fullmatch(r'\d{4}', token):
			return ('year', int(token))
		if re.fullmatch(r'\d{4}-\d{2}', token):
			year, month = token.split('-')
			return ('month', int(year), int(month))
		d = parse_date(token)
		if d:
			return ('date', d)
		return None

	def apply_exact(qs, parsed):
		kind = parsed[0]
		if kind == 'year':
			return qs.filter(**{f'{field}__year': parsed[1]})
		if kind == 'month':
			return qs.filter(**{f'{field}__year': parsed[1], f'{field}__month': parsed[2]})
		return qs.filter(**{f'{field}__date': parsed[1]})

	def apply_exclude_exact(qs, parsed):
		kind = parsed[0]
		if kind == 'year':
			return qs.exclude(**{f'{field}__year': parsed[1]})
		if kind == 'month':
			return qs.exclude(**{f'{field}__year': parsed[1], f'{field}__month': parsed[2]})
		return qs.exclude(**{f'{field}__date': parsed[1]})

	def token_as_date(parsed):
		kind = parsed[0]
		if kind == 'date':
			return parsed[1]
		return None

	op_match = re.match(r'^(<=|>=|<|>|=|!=)\s*(.+)$', raw)
	range_match = re.match(r'^(.+?)\.\.(.+)$', raw)
	if not range_match:
		range_match = re.match(r'^(\d{4}-\d{2}-\d{2})\s*-\s*(\d{4}-\d{2}-\d{2})$', raw)

	if op_match:
		op, rhs_raw = op_match.groups()
		parsed = parse_token(rhs_raw)
		if not parsed:
			return queryset
		if op == '=':
			return apply_exact(queryset, parsed)
		if op == '!=':
			return apply_exclude_exact(queryset, parsed)

		rhs_date = token_as_date(parsed)
		if rhs_date is None:
			return queryset

		lookup_map = {
			'<=': 'lte',
			'>=': 'gte',
			'<': 'lt',
			'>': 'gt',
		}
		lookup = lookup_map.get(op)
		if not lookup:
			return queryset
		return queryset.filter(**{f'{field}__date__{lookup}': rhs_date})

	if range_match:
		left_raw, right_raw = range_match.groups()
		left_parsed = parse_token(left_raw)
		right_parsed = parse_token(right_raw)
		left_date = token_as_date(left_parsed) if left_parsed else None
		right_date = token_as_date(right_parsed) if right_parsed else None
		if left_date is None or right_date is None:
			return queryset
		low = min(left_date, right_date)
		high = max(left_date, right_date)
		return queryset.filter(**{f'{field}__date__gte': low, f'{field}__date__lte': high})

	parsed = parse_token(raw)
	if not parsed:
		# Fallback to substring matching (similar to previous frontend behavior),
		# e.g. input "22" can match datetimes whose rendered value contains "22".
		alias = f'_{field}_str'
		return queryset.annotate(**{alias: Cast(field, CharField())}).filter(**{f'{alias}__icontains': raw})
	return apply_exact(queryset, parsed)

def list(request):
	orders = Order.objects.all()
	needs_distinct = False
	urgent_only = _clean_param(request, 'urgent_only') == '1'
	duplicate_only = _clean_param(request, 'duplicate_only') == '1'
	special_fees_only = _clean_param(request, 'special_fees_only') == '1'

	for field in ['reference', 'contact_name', 'phone', 'suburb', 'postcode', 'state', 'special_fees', 'customer_notes', 'notes']:
		orders = _apply_text_filter(orders, field, _clean_param(request, field))

	orders = _apply_date_filter(orders, 'date', _clean_param(request, 'date'))
	orders = _apply_number_filter(orders, 'total', _clean_param(request, 'total'))
	orders = _apply_number_filter(orders, 'shipping', _clean_param(request, 'shipping'))

	woo_status = _clean_param(request, 'woo_status')
	if woo_status:
		orders = orders.filter(woo_status=woo_status)

	route_record = _clean_param(request, 'route_record')
	if route_record:
		orders = orders.filter(route_record=route_record)

	status_param_present = 'status' in request.GET
	status = _clean_param(request, 'status')
	if status_param_present:
		if status == 'Open':
			orders = orders.exclude(status__in=['Completed', 'Cancelled'])
		elif status:
			orders = orders.filter(status=status)
	else:
		orders = orders.exclude(status__in=['Completed', 'Cancelled'])

	products = _clean_param(request, 'products')
	if products:
		orders = orders.filter(
			Q(lines__raw_sku__icontains=products) |
			Q(lines__product__sku__icontains=products) |
			Q(lines__product__barcode__icontains=products) |
			Q(lines__product__name_cn__icontains=products) |
			Q(lines__product__name_en__icontains=products)
		)
		needs_distinct = True

	if needs_distinct:
		orders = orders.distinct()

	if urgent_only:
		orders = orders.filter(urgent=True)

	orders = orders.prefetch_related('lines').order_by('-date')

	def _normalize_duplicate_key(contact_name, phone):
		name = ' '.join((contact_name or '').strip().lower().split())
		phone_digits = ''.join(ch for ch in str(phone or '') if ch.isdigit())
		phone_value = phone_digits[-9:] if len(phone_digits) >= 9 else phone_digits
		return name, phone_value

	duplicate_counter = {}
	for contact_name, phone in (
		Order.objects
		.exclude(status='Cancelled')
		.exclude(contact_name__isnull=True)
		.exclude(phone__isnull=True)
		.exclude(contact_name='')
		.exclude(phone='')
		.values_list('contact_name', 'phone')
	):
		key = _normalize_duplicate_key(contact_name, phone)
		if not key[0] or not key[1]:
			continue
		duplicate_counter[key] = duplicate_counter.get(key, 0) + 1

	duplicate_keys = {key for key, count in duplicate_counter.items() if count > 1}
	orders = builtins.list(orders)
	for order in orders:
		if order.status == 'Cancelled':
			order.has_same_contact_phone_orders = False
			continue
		order.has_same_contact_phone_orders = _normalize_duplicate_key(order.contact_name, order.phone) in duplicate_keys

	if duplicate_only:
		orders = [order for order in orders if order.has_same_contact_phone_orders]

	if special_fees_only:
		orders = [order for order in orders if (order.special_fees or '').strip()]

	def _build_toggle_url(param_name, enabled):
		params = request.GET.copy()
		if enabled:
			params[param_name] = '1'
			if param_name == 'urgent_only':
				params.pop('duplicate_only', None)
				params.pop('special_fees_only', None)
			elif param_name == 'duplicate_only':
				params.pop('urgent_only', None)
				params.pop('special_fees_only', None)
			elif param_name == 'special_fees_only':
				params.pop('urgent_only', None)
				params.pop('duplicate_only', None)
		else:
			params.pop(param_name, None)
		query = params.urlencode()
		return f'{request.path}?{query}' if query else request.path

	return render(request, 'orders/list.html', {
		'orders': orders,
		'woo_statuses': ORDER_WOO_STATUS,
		'statuses': ORDER_STATUS,
		'route_records': ORDER_ROUTE_RECORD,
		'urgent_only': urgent_only,
		'duplicate_only': duplicate_only,
		'special_fees_only': special_fees_only,
		'urgent_toggle_url': _build_toggle_url('urgent_only', not urgent_only),
		'duplicate_toggle_url': _build_toggle_url('duplicate_only', not duplicate_only),
		'special_fees_toggle_url': _build_toggle_url('special_fees_only', not special_fees_only),
	})

@csrf_exempt
def create_order(request):
	if request.method != 'POST':
		return JsonResponse({'success': False, 'error': 'Invalid method'})

	try:
		data = json.loads(request.body)
	except json.JSONDecodeError:
		logger.warning('create_order failed: invalid JSON payload')
		return JsonResponse({'success': False, 'error': 'Invalid JSON payload'})

	try:
		reference = data.get('reference')
		contact_name = data.get('contact_name')
		phone = data.get('phone')
		email = data.get('email')
		address = data.get('address')
		suburb = data.get('suburb')
		postcode = data.get('postcode')
		state = data.get('state')
		route_record = data.get('route_record')
		notes = data.get('notes')
		customer_notes = data.get('customer_notes', '')
		total = _parse_decimal_value(data.get('total'))
		shipping = _parse_decimal_value(data.get('shipping'))
		special_fees = (data.get('special_fees') or '').strip()
		status = data.get('status')
		urgent = parse_bool(data.get('urgent', False))
		date = data.get('date')
		tracking_number = data.get('tracking_number', '')
		delivery_date = data.get('delivery_date', '')
		parcel_attributes = normalize_parcel_attributes(data.get('parcel_attributes'))
		if delivery_date == '':
			delivery_date = '1970-01-01'
		products = data.get('products', [])

		meta = {}
		if parcel_attributes:
			meta['parcel_attributes'] = parcel_attributes
		order = Order.objects.create(reference=reference, date=date, contact_name=contact_name, phone=phone, email=email, 
			address=address, suburb=suburb, postcode=postcode, state=state, route_record=route_record, notes=notes, customer_notes=customer_notes, status=status, total=total, shipping=shipping, special_fees=special_fees, urgent=urgent, tracking_number=tracking_number, delivery_date=delivery_date, meta=meta or None)

		for item in products:
			product = None
			product_id = item.get('product_id')
			if product_id:
				product = Product.objects.filter(id=product_id).first()
			OrderLine.objects.create(
				order=order,
				product=product,
				raw_sku=item.get('sku', ''),
				quantity=item.get('quantity', 0)
			)

		Stock.recalculate_all()
		if order.reference == '':
			push_order_to_wc(order.id)
		return JsonResponse({'success': True})
	except Exception as e:
		safe_payload = data.copy() if isinstance(data, dict) else {}
		if isinstance(safe_payload, dict):
			for key in ['phone', 'email', 'address']:
				if key in safe_payload and safe_payload.get(key):
					safe_payload[key] = '***'
		logger.exception('create_order failed. payload=%s', safe_payload)
		error_message = str(e) if settings.DEBUG else 'Create order failed'
		return JsonResponse({'success': False, 'error': error_message}, status=500)

@csrf_exempt
def delete_order(request, id):
	if request.method == 'POST':
		try:
			order = Order.objects.filter(id=id).first()
			if not order:
				return JsonResponse({'success': False, 'error': '未找到订单'})

			order.delete()
			Stock.recalculate_all()
			return JsonResponse({'success': True})
		
		except json.JSONDecodeError:
			return JsonResponse({'success': False, 'error': '无效的 JSON 格式'})
		except Exception as e:
			return JsonResponse({'success': False, 'error': str(e)})
	
	return JsonResponse({'success': False, 'error': '仅支持 POST 请求'})

@csrf_exempt
def order_detail(request, id):
	if request.method == 'GET':
		try:
			order = Order.objects.get(id=id)
			lines = order.lines.all()

			if order.delivery_date and order.delivery_date.strftime('%Y-%m-%d') == '1970-01-01':
				order.delivery_date = None

			data = {
				'id': order.id,
				'reference': order.reference,
				'date': str(order.date),
				'contact_name': order.contact_name,
				'phone': order.phone,
				'email': order.email,
				'address': order.address,
				'suburb': order.suburb,
				'postcode': order.postcode,
				'state': order.state,
				'route_record': order.route_record,
				'notes': order.notes,
				'customer_notes': order.customer_notes,
				'total': str(order.total) if order.total is not None else '',
				'shipping': str(order.shipping) if order.shipping is not None else '',
				'special_fees': order.special_fees,
				'status': order.status,
				'urgent': order.urgent,
				'tracking_number': order.tracking_number,
				'delivery_date': order.delivery_date,
				'parcel_attributes': normalize_parcel_attributes((order.meta or {}).get('parcel_attributes')) or build_parcel_attributes(order),
				'products': [
					{
						'product_id': line.product.id if line.product else '',
						'name': line.product.name_cn if line.product else '',
						'barcode': line.product.barcode if line.product else '',
						'sku': line.product.sku if line.product else line.raw_sku,
						'quantity': line.quantity,
					} for line in lines
				]
			}
			return JsonResponse({'success': True, 'order': data})
		except Exception as e:
			return JsonResponse({'success': False, 'error': str(e)})
	
	return JsonResponse({'success': False, 'error': '仅支持 GET 请求'})

@csrf_exempt
def update_order(request, id):
	if request.method == 'POST':
		try:
			data = json.loads(request.body)

			reference = data.get('reference')
			date = data.get('date')
			contact_name = data.get('contact_name')
			phone = data.get('phone')
			email = data.get('email')
			address = data.get('address')
			suburb = data.get('suburb')
			postcode = data.get('postcode')
			state = data.get('state')
			route_record = data.get('route_record')
			notes = data.get('notes')
			customer_notes = data.get('customer_notes', '')
			total = _parse_decimal_value(data.get('total'))
			shipping = _parse_decimal_value(data.get('shipping'))
			special_fees = (data.get('special_fees') or '').strip()
			status = data.get('status')
			urgent = parse_bool(data.get('urgent', False))
			tracking_number = data.get('tracking_number', '')
			delivery_date = data.get('delivery_date', '')
			parcel_attributes = normalize_parcel_attributes(data.get('parcel_attributes'))
			if delivery_date == '':
				delivery_date = '1970-01-01'
			if (status or '').strip().lower() == 'completed' and _is_empty_delivery_date(delivery_date):
				delivery_date = timezone.localdate().isoformat()
			products = data.get('products', [])

			order = Order.objects.get(id=id)
			order.reference = reference
			order.date = date
			order.contact_name = contact_name
			order.phone = phone
			order.email = email
			order.address = address
			order.suburb = suburb
			order.postcode = postcode
			order.state = state
			order.route_record = route_record
			order.notes = notes
			order.customer_notes = customer_notes
			order.total = total
			order.shipping = shipping
			order.special_fees = special_fees
			order.status = status
			order.urgent = urgent
			order.tracking_number = tracking_number
			order.delivery_date = delivery_date
			meta = order.meta if isinstance(order.meta, dict) else {}
			if parcel_attributes:
				meta['parcel_attributes'] = parcel_attributes
			else:
				meta.pop('parcel_attributes', None)
			order.meta = meta or None
			order.save()

			# 删除旧的明细行
			order.lines.all().delete()

			# 添加新的明细行
			for item in products:
				product = None
				product_id = item.get('product_id')
				quantity = item.get('quantity', 0)

				if product_id:
					product = Product.objects.filter(id=product_id).first()

				OrderLine.objects.create(
					order=order,
					product=product,
					raw_sku=item.get('sku', ''),
					quantity=quantity
				)

			Stock.recalculate_all()
			if (order.status or '').strip().lower() == 'completed':
				sync_woo_order_completed(order)
			return JsonResponse({'success': True})

		except Order.DoesNotExist:
			return JsonResponse({'success': False, 'error': '未找到订单'})
		except json.JSONDecodeError:
			return JsonResponse({'success': False, 'error': '无效的 JSON 格式'})
		except Exception as e:
			return JsonResponse({'success': False, 'error': str(e)})

	return JsonResponse({'success': False, 'error': '仅支持 POST 请求'})

@csrf_exempt
def batch_update_order(request):
	if request.method != 'POST':
		return JsonResponse({'success': False, 'error': '仅支持 POST 请求'})

	try:
		data = json.loads(request.body or '{}')
	except json.JSONDecodeError:
		return JsonResponse({'success': False, 'error': '无效的 JSON 格式'})

	raw_ids = data.get('ids') or []
	status = (data.get('status') or '').strip()

	if not isinstance(raw_ids, builtins.list):
		return JsonResponse({'success': False, 'error': '订单ID格式错误'}, status=400)

	id_list = []
	for item in raw_ids:
		try:
			id_list.append(int(item))
		except (TypeError, ValueError):
			continue
	id_list = tuple(dict.fromkeys(id_list))

	if not id_list:
		return JsonResponse({'success': False, 'error': '请先勾选要更新的订单'}, status=400)

	valid_status_values = {value for value, _ in ORDER_STATUS}
	if status not in valid_status_values:
		return JsonResponse({'success': False, 'error': '请选择有效的订单状态'}, status=400)

	orders = Order.objects.filter(id__in=id_list)
	if not orders:
		return JsonResponse({'success': False, 'error': '未找到可更新的订单'}, status=404)

	updated_count = 0
	for order in orders:
		if order.status == status:
			continue
		order.status = status
		update_fields = ['status']
		if (status or '').strip().lower() == 'completed' and _is_empty_delivery_date(order.delivery_date):
			order.delivery_date = timezone.localdate()
			update_fields.append('delivery_date')
		order.save(update_fields=update_fields)
		updated_count += 1
		if (order.status or '').strip().lower() == 'completed':
			sync_woo_order_completed(order)

	Stock.recalculate_all()

	return JsonResponse({
		'success': True,
		'updated_count': updated_count,
		'matched_count': len(orders),
	})

@csrf_exempt
def _export_orders_excel_response(orders, filename):
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = 'Order Export'

	headers = [
		'Order No',
		'Order Date',
		'Status',
		'Woo Status',
		'Customer',
		'Phone',
		'Email',
		'Address',
		'Suburb',
		'Postcode',
		'State',
		'Products',
		'Total',
		'Shipping',
		'Special Fee',
		'Tracking No',
		'Delivery Date',
		'Route Record',
		'Customer Notes',
		'Notes',
	]
	ws.append(headers)
	woo_status_map = dict(ORDER_WOO_STATUS)

	for order in orders:
		order_dt = order.date
		if order_dt and timezone.is_aware(order_dt):
			order_dt = timezone.localtime(order_dt)
		delivery_date_str = ''
		if order.delivery_date and order.delivery_date.strftime('%Y-%m-%d') != '1970-01-01':
			delivery_date_str = order.delivery_date.strftime('%Y-%m-%d')
		product_lines = '\n'.join([
			f'{(line.raw_sku or (line.product.sku if (line.product and line.product.sku) else ""))} x {line.quantity}'
			for line in order.lines.all()
		])

		ws.append([
			order.reference or '',
			order_dt.strftime('%Y-%m-%d %H:%M:%S') if order_dt else '',
			order.status or '',
			woo_status_map.get(order.woo_status, order.woo_status or ''),
			order.contact_name or '',
			order.phone or '',
			order.email or '',
			order.address or '',
			order.suburb or '',
			order.postcode or '',
			order.state or '',
			product_lines,
			str(order.total or ''),
			str(order.shipping or ''),
			order.special_fees or '',
			order.tracking_number or '',
			delivery_date_str,
			order.route_record or '',
			order.customer_notes or '',
			order.notes or '',
		])

	header_font = Font(name='Arial', size=11, bold=True)
	body_font = Font(name='Arial', size=10)
	alignment = Alignment(vertical='top', wrap_text=True)
	header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

	for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
		for cell in row:
			cell.font = header_font if row_idx == 1 else body_font
			cell.alignment = header_alignment if row_idx == 1 else alignment

	ws.freeze_panes = 'A2'

	def text_width(val):
		return sum(2 if ord(c) > 127 else 1 for c in str(val))

	for col in ws.columns:
		max_length = 0
		col_letter = get_column_letter(col[0].column)
		for cell in col:
			if cell.value:
				max_length = max(max_length, text_width(cell.value))
		ws.column_dimensions[col_letter].width = max(12, min((max_length + 2) * 1.1, 60))

	for i in range(2, ws.max_row + 1):
		ws.row_dimensions[i].height = 36

	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename="{filename}"'
	wb.save(response)
	return response


@csrf_exempt
def export_orders(request):
	ids = (request.GET.get('ids') or '').strip()
	if ids:
		export_type = (request.GET.get('export_type') or '').strip().lower()
		if export_type == 'orders':
			id_list = [int(x) for x in ids.split(',') if x.isdigit()]
			if not id_list:
				return JsonResponse({'success': False, 'error': '请先选择要导出的订单'}, status=400)

			orders = (
				Order.objects
				.filter(id__in=id_list)
				.order_by('date', 'id')
				.prefetch_related('lines')
			)
			filename = f'orders_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx'
			return _export_orders_excel_response(orders, filename)
		return export_orders_delivery(request)

	start_date_raw = (request.GET.get('start_date') or '').strip()
	end_date_raw = (request.GET.get('end_date') or '').strip()
	start_date = parse_date(start_date_raw) if start_date_raw else None
	end_date = parse_date(end_date_raw) if end_date_raw else None

	if not start_date or not end_date:
		return JsonResponse({'success': False, 'error': '请提供开始日期和结束日期'}, status=400)

	if start_date > end_date:
		return JsonResponse({'success': False, 'error': '开始日期不能晚于结束日期'}, status=400)

	orders = (
		Order.objects
		.filter(date__date__gte=start_date, date__date__lte=end_date)
		.order_by('date', 'id')
		.prefetch_related('lines')
	)

	filename = f'orders_{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}.xlsx'
	return _export_orders_excel_response(orders, filename)

@csrf_exempt
def export_orders_delivery(request):
	ids = request.GET.get('ids', '')
	id_list = [int(x) for x in ids.split(',') if x.isdigit()]

	if not id_list:
		return JsonResponse({'success': False, 'error': '请先选择要导出的订单'}, status=400)

	orders = Order.objects.filter(id__in=id_list).prefetch_related('lines')

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = 'Orders'

	font = Font(name='Arial', size=14)
	alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

	from openpyxl.styles import Border, Side
	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)

	# 行高
	for i in range(1, len(orders) + 10):
		ws.row_dimensions[i].height = 74.25

	# 抬头日期（次日）
	header_date = (timezone.localtime() + timedelta(days=1)).strftime('%m-%d-%Y')
	ws.append(['DELIVERY LIST'] + [''] * 11 + ['DATE', header_date])

	# 第二行：表头
	headers = [
		'NUMBER',
		'ITEM DESCRIPTION',
		'DELIVERY INSTRUCTION',
		'PACKAGE',
		'CUSTOMER NAME',
		'CONTACT NO.',
		'ADDRESS',
		'SUBURB',
		'POST CODE',
		'STATE',
		'Invoice No.',
		'NOTE',
		'SIGNATURE OF RECEIPIENT',
		'DATE',
	]
	ws.append(headers)

	# 数据行
	for order in orders:
		product_lines = '\n'.join([
			f'{line.raw_sku} × {line.quantity}'
			for line in order.lines.all()
		])

		packages = []
		for line in order.lines.all():
			product = line.product
			if not product:
				packages.append(f'{line.raw_sku} x {line.quantity}')
				continue

			bom_items = product.bom_items.select_related('component').all()
			if bom_items.exists():
				for bom in bom_items:
					packages.append(f'{bom.component.sku} × {bom.quantity * line.quantity}')
			else:
				packages.append(f'{product.sku} × {line.quantity}')
		packages = '\n'.join(packages)

		ws.append([
			order.reference,
			product_lines,
			'',
			packages,
			order.contact_name or '',
			order.phone or '',
			order.address or '',
			order.suburb or '',
			order.postcode or '',
			order.state or '',
			'',
			order.notes or '',
			'',
			header_date
		])

	# 样式 + 边框
	for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
		for cell in row:
			cell.font = font
			cell.alignment = alignment
			cell.border = thin_border

	# 列宽自适应（考虑中文宽度）
	def text_width(val):
		return sum(2 if ord(c) > 127 else 1 for c in str(val))

	for col in ws.columns:
		max_length = 0
		col_letter = get_column_letter(col[0].column)
		for cell in col:
			if cell.value:
				val_length = text_width(cell.value)
				if val_length > max_length:
					max_length = val_length
		adjusted_width = max(15, min((max_length + 2) * 1.2, 50))
		ws.column_dimensions[col_letter].width = adjusted_width

	# 页面设置
	ws.sheet_properties.pageSetUpPr.fitToPage = True
	ws.page_setup.fitToHeight = False
	ws.page_setup.fitToWidth = 1
	ws.page_setup.scale = 55
	ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

	filename = f'delivery_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx'
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = f'attachment; filename="{filename}"'
	wb.save(response)

	return response

def invoice_page(request, id):
	order = Order.objects.prefetch_related('lines__product').filter(id=id).first()
	if not order:
		return render(request, 'errors/404.html', status=404)

	meta = order.meta if isinstance(order.meta, dict) else {}
	meta_line_items = []
	if meta:
		meta_line_items = meta.get('line_items', []) or []

	meta_line_map = {}
	for item in meta_line_items:
		if not isinstance(item, dict):
			continue
		for key in [str(item.get('sku') or '').strip(), str(item.get('name') or '').strip()]:
			if key and key not in meta_line_map:
				meta_line_map[key] = item

	def _to_decimal(raw, default=Decimal('0')):
		try:
			if raw in [None, '']:
				return default
			return Decimal(str(raw))
		except (InvalidOperation, TypeError, ValueError):
			return default

	def _format_compact_number(raw):
		if raw in [None, '']:
			return ''
		try:
			d = Decimal(str(raw))
		except (InvalidOperation, TypeError, ValueError):
			return str(raw)
		if d == d.to_integral():
			return str(d.quantize(Decimal('1')))
		return format(d.normalize(), 'f').rstrip('0').rstrip('.')

	def _clean_display_text(value):
		if isinstance(value, (dict, builtins.list)):
			return ''
		text = str(value or '').strip()
		text = unescape(text)
		text = py_re.sub(r'<[^>]+>', '', text)
		return text.strip()

	def _extract_line_meta_details(meta_item):
		rows = []
		for md in (meta_item.get('meta_data') or []):
			if not isinstance(md, dict):
				continue
			key = str(md.get('key') or '').strip()
			if not key or key.startswith('_') or key in ['_reduced_stock']:
				continue
			label = _clean_display_text(md.get('display_key') or key)
			value = _clean_display_text(md.get('display_value') or md.get('value'))
			if not label or not value:
				continue
			rows.append({'label': label, 'value': value})
		return rows

	AU_STATE_NAMES = {
		'NSW': 'New South Wales',
		'VIC': 'Victoria',
		'QLD': 'Queensland',
		'SA': 'South Australia',
		'WA': 'Western Australia',
		'TAS': 'Tasmania',
		'ACT': 'Australian Capital Territory',
		'NT': 'Northern Territory',
	}

	line_rows = []
	prices_include_tax = bool(meta.get('prices_include_tax')) if meta else False
	for line in order.lines.all():
		display_name = line.display_name_en() if hasattr(line, 'display_name_en') else (line.raw_sku or '')
		sku = (line.product.sku if line.product else line.raw_sku) or ''
		meta_item = meta_line_map.get(str(sku).strip()) or meta_line_map.get(display_name.strip()) or {}
		line_subtotal = _to_decimal(meta_item.get('subtotal'), default=None)
		line_subtotal_tax = _to_decimal(meta_item.get('subtotal_tax'), default=Decimal('0')) if line_subtotal is not None else Decimal('0')
		line_total_value = ''
		if line_subtotal is not None:
			line_total_value = line_subtotal + (line_subtotal_tax if prices_include_tax else Decimal('0'))
		elif meta_item.get('total') not in [None, '']:
			line_total_base = _to_decimal(meta_item.get('total'))
			line_total_tax = _to_decimal(meta_item.get('total_tax'))
			line_total_value = line_total_base + (line_total_tax if prices_include_tax else Decimal('0'))
		product = line.product
		weight = ''
		size = ''
		if product:
			if product.weight is not None:
				weight = f"{_format_compact_number(product.weight)}kg"
			length = (product.package_length or '').strip()
			width = (product.package_width or '').strip()
			height = (product.package_height or '').strip()
			dims = [_format_compact_number(v) for v in [length, width, height] if v]
			if dims:
				size = ' x '.join(dims)
		line_rows.append({
			'name': display_name,
			'sku': sku,
			'qty': line.quantity,
			'line_total': line_total_value,
			'weight': weight,
			'size': size,
			'meta_details': _extract_line_meta_details(meta_item),
		})

	special_fee_lines = []
	for raw in (order.special_fees or '').splitlines():
		raw = (raw or '').strip()
		if not raw:
			continue
		if ':' in raw:
			name, value = raw.split(':', 1)
			special_fee_lines.append({'name': name.strip(), 'value': value.strip()})
		else:
			special_fee_lines.append({'name': raw, 'value': ''})

	shipping_method = ''
	for shipping_line in (meta.get('shipping_lines') or []):
		if isinstance(shipping_line, dict):
			shipping_method = (shipping_line.get('method_title') or '').strip()
			if shipping_method:
				break

	payment_method_label = (meta.get('payment_method_title') or meta.get('payment_method') or '') if meta else ''
	woo_order_number = str(meta.get('number') or order.reference or '') if meta else str(order.reference or '')
	order_date_display = ''
	if order.date:
		try:
			order_date_display = timezone.localtime(order.date).strftime('%B %d, %Y').replace(' 0', ' ')
		except Exception:
			order_date_display = order.date.strftime('%B %d, %Y').replace(' 0', ' ')

	total_amount = _to_decimal(meta.get('total')) if meta else _to_decimal(order.total)
	shipping_amount = _to_decimal(meta.get('shipping_total')) if meta else _to_decimal(order.shipping)
	if not meta:
		shipping_amount = _to_decimal(order.shipping)
	shipping_tax_amount = _to_decimal(meta.get('shipping_tax')) if meta else Decimal('0')
	discount_amount = (_to_decimal(meta.get('discount_total')) + _to_decimal(meta.get('discount_tax'))) if meta else Decimal('0')
	shipping_display_amount = shipping_amount + (shipping_tax_amount if prices_include_tax else Decimal('0'))
	subtotal_amount = total_amount + discount_amount - shipping_display_amount
	if subtotal_amount < 0:
		subtotal_amount = Decimal('0')

	shipping_display = ''
	if shipping_display_amount == 0 and shipping_method:
		shipping_display = shipping_method
	else:
		shipping_display = f'${shipping_display_amount}'
		if shipping_method:
			shipping_display = f'{shipping_display} via {shipping_method}'

	gst_amount = None
	if meta:
		try:
			raw_tax = meta.get('total_tax')
			if raw_tax not in [None, '']:
				gst_amount = Decimal(str(raw_tax))
		except (InvalidOperation, TypeError, ValueError):
			gst_amount = None

	billing = meta.get('billing') if isinstance(meta.get('billing'), dict) else {}
	bill_name = ' '.join([str(billing.get('first_name') or '').strip(), str(billing.get('last_name') or '').strip()]).strip() or (order.contact_name or '')
	bill_addr_parts = [str(billing.get('address_1') or '').strip(), str(billing.get('address_2') or '').strip()]
	bill_address = ', '.join([p for p in bill_addr_parts if p]) or (order.address or '')
	bill_city = str(billing.get('city') or order.suburb or '').strip()
	bill_state_raw = str(billing.get('state') or order.state or '').strip().upper()
	bill_state = AU_STATE_NAMES.get(bill_state_raw, bill_state_raw)
	bill_postcode = str(billing.get('postcode') or order.postcode or '').strip()
	bill_city_state_postcode = ' '.join([p for p in [bill_city, bill_state, bill_postcode] if p]).strip()

	return render(request, 'orders/invoice.html', {
		'order': order,
		'meta': meta,
		'line_rows': line_rows,
		'special_fee_lines': special_fee_lines,
		'invoice_company': {
			'name': 'TOPONE SPORTS PTY LTD',
			'abn': '84 679 476 478',
			'address_1': '181-185 Parramatta Road Granville',
			'address_2': 'NSW 2142',
			'phone': '(02) 6188 5799',
		},
		'woo_order_number': woo_order_number,
		'order_date_display': order_date_display,
		'payment_method_label': payment_method_label,
		'shipping_method_label': shipping_method,
		'shipping_display': shipping_display,
		'subtotal_amount': subtotal_amount,
		'shipping_amount': shipping_amount,
		'discount_amount': discount_amount,
		'has_discount_amount': discount_amount > 0,
		'total_amount': total_amount,
		'gst_amount': gst_amount,
		'has_gst_amount': gst_amount is not None,
		'bill_name': bill_name,
		'bill_address': bill_address,
		'bill_city_state_postcode': bill_city_state_postcode,
		'print_mode': request.GET.get('print') == '1',
	})

@csrf_exempt
def shipping_quotes(request, id):
	try:
		result = get_shipping_quote(id)
		if not result:
			return JsonResponse({'success': False, 'error': '未获取到任何报价'})
		return JsonResponse({
			'success': True,
			'quotes': result['all'],
			'best': result['best'],
			'parcel_attributes': result['parcel_attributes'],
		})
	except Exception as e:
		return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
def sync_orders(request):
	if request.method == 'POST':
		try:
			sync_wc_orders()
			return JsonResponse({'success': True})
		except Exception as e:
			return JsonResponse({'success': False, 'error': str(e)})
	
	return JsonResponse({'success': False, 'error': '仅支持 POST 请求'})
